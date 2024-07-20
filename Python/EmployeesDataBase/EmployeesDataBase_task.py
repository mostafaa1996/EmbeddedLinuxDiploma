from tkinter import filedialog as fd
from tkinter.filedialog import asksaveasfile
import openpyxl as xl
import pandas as pd
from openpyxl.styles import Alignment
import os 

Employee_Info = {"Name": "" , "Title" : "" , "ID" : ""   , "Salary" : "" }
Employees_list_ToBe_write = []
Employees_list_ToBe_Read = []
openedExcelSheet = None
fileName = ""

def Generate_Excel(name=None , Data = None):
    DataBase = pd.DataFrame(Data)
    if(name == None):
        f = asksaveasfile(initialfile = 'Untitled.xlsx',
            defaultextension=".xlsx",filetypes=[("All Files","*.*"),("Excel Format","*.xlsx")])
        print(f.name)
        DataBase.to_excel(f.name, sheet_name='Sheet1', index=False, startrow=1)
        FormatEXcel(f.name)
        f.close()
    else:
        DataBase.to_excel(name, sheet_name='Sheet1', index=False, startrow=1)
        FormatEXcel(name)
        
    
def Read_Excel():
    global Employees_list_ToBe_Read
    global fileName
    file = fd.askopenfile()
    fileName = file.name
    book = xl.load_workbook(file.name)
    sheet = book.active
    rows_data = []

    # Get the headers from the first row
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=2, max_row=2))]

    # Iterate over the remaining rows and convert each row to a dictionary
    for row in sheet.iter_rows(min_row=3, values_only=True):
        row_data = dict(zip(headers, row))
        rows_data.append(row_data.copy())
    # for i in rows_data :
    #     i.pop("Attributes")
    print(rows_data)
    file.close()
    Employees_list_ToBe_Read = rows_data.copy()

def ModifyItemsIn_Excel(item):
    global Employees_list_ToBe_Read
    global fileName
    Read_Excel()
    RequiredIndex = 0
    for i in Employees_list_ToBe_Read:
        if(i["Name"].lower() == item["Name"].lower() or i["ID"].lower() == item["ID"].lower()):
            RequiredIndex = Employees_list_ToBe_Read.index(i)
            
            Employees_list_ToBe_Read.remove(i)
            Employees_list_ToBe_Read.insert(RequiredIndex , item)
            break
    Employees_list_ToBe_write = Employees_list_ToBe_Read.copy()
    Generate_Excel(fileName , Employees_list_ToBe_write)

def ShowItemData():
    pass

def AddItemToexisting_Excel():
    global Employee_Info
    Employees_list_ToBe_write.append(Employee_Info.copy())

def Open_Excel():
    global fileName
    file = fd.askopenfile()
    fileName = file.name
    file.close()
    
def appendToOpenedExcel():
    global fileName
    DataBase = pd.DataFrame(Employees_list_ToBe_write)
    book = xl.load_workbook(fileName)
    sheet = book.active
    # Append new rows
    for row in DataBase.itertuples(index=False, name=None):
        sheet.append(row)

    # Save the workbook
    book.save(fileName)

    
def RemoveItem(item):
    global Employees_list_ToBe_write
    file = fd.askopenfile()
    NameOfFile = file.name
    file.close()
    df = pd.read_excel(NameOfFile, sheet_name='Sheet1')
    Dict = df.to_dict()
    RequiredIndex = 0
    found = False
    for i in Dict.values() :
        if(item["ID"]!= ""):
          if(i[0] == "ID"):
             for ind in range(1,len(i)):
                 if i[ind] == item["ID"]:
                     RequiredIndex = ind
                     break
        elif(item["Name"]!= ""):
          if(i[0] == "Name"):
             for ind in range(1,len(i)):
                 if i[ind].lower() == item["Name"].lower():
                     RequiredIndex = ind
                     found = True
                     break
          if found == True:
            break
    df = df.drop(index=RequiredIndex)
    Employees_list_ToBe_write = []
    for i in range(1,len(df.values)):
        loc_info = {}
        loc_info[df.values[0][0]] = df.values[i][0]
        loc_info[df.values[0][1]] = df.values[i][1]
        loc_info[df.values[0][2]] = df.values[i][2]
        loc_info[df.values[0][3]] = df.values[i][3]
        Employees_list_ToBe_write.append(loc_info.copy())
    Generate_Excel(NameOfFile , Employees_list_ToBe_write)

def FormatEXcel(name):
    # Load the workbook and select the sheet
    wb = xl.load_workbook(name)
    ws = wb['Sheet1']
    
    # Insert the main header and merge cells
    main_header = "Employees Information"
    ws['A1'] = main_header
    ws.merge_cells('A1:D1')

    # Center align the main header
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    # Save the workbook
    wb.save(name)